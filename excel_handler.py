"""Leitura do Excel e coloração de células para tracking."""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill

from config import EXCEL_PATH, EXCEL_COLUMNS, MAX_ATTRIBUTES, DOCUMENTS_DIR
from logger import log

# Cores para pintar linhas (compatível com o fluxo do André)
FILL_GREEN = PatternFill(start_color="00FF00", fill_type="solid")   # sucesso
FILL_RED = PatternFill(start_color="FF0000", fill_type="solid")     # erro
FILL_ORANGE = PatternFill(start_color="FFA500", fill_type="solid")  # duplicidade
FILL_YELLOW = PatternFill(start_color="FFFF00", fill_type="solid")  # needs_review


def load_excel(path: Path | None = None) -> tuple[openpyxl.Workbook, list[dict[str, Any]]]:
    """Lê a planilha e retorna (workbook, lista de dicts por linha).

    Cada dict tem as chaves definidas em EXCEL_COLUMNS + Atrib_N_Valor.
    Também inclui '_row' com o número da linha no Excel (para colorir).
    """
    path = path or EXCEL_PATH
    log.info(f"Abrindo Excel: {path}")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Encontrar header row (primeira linha com dados nas colunas esperadas)
    headers = {}
    header_row = 1
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        row_vals = [cell.value for cell in ws[row_idx]]
        row_vals_upper = [str(v).upper() if v else "" for v in row_vals]
        if "SIN" in row_vals_upper or any(
            str(v).upper() in row_vals_upper for v in EXCEL_COLUMNS.values()
        ):
            # Mapa case-insensitive: header_upper -> col_idx
            raw_headers = {
                ws.cell(row=row_idx, column=col).value: col
                for col in range(1, ws.max_column + 1)
                if ws.cell(row=row_idx, column=col).value
            }
            # Mapa normalizado para lookup case-insensitive
            headers_upper = {str(k).upper(): col for k, col in raw_headers.items()}
            headers = raw_headers
            headers["_upper"] = headers_upper
            header_row = row_idx
            break

    if not headers:
        raise ValueError("Não foi possível encontrar cabeçalho no Excel")

    log.info(f"Cabeçalho encontrado na linha {header_row}: {list(headers.keys())[:10]}...")

    items = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data = {"_row": row_idx}

        # Campos padrão (case-insensitive + aliases)
        _upper = headers.get("_upper", {})
        for key, col_name in EXCEL_COLUMNS.items():
            col_idx = headers.get(col_name) or _upper.get(str(col_name).upper())
            row_data[key] = ws.cell(row=row_idx, column=col_idx).value if col_idx else None

        # Aliases para colunas com nomes diferentes entre planilhas
        _aliases = {
            "codigo_60": ["CÓDIGO 60", "CODIGO 60", "COD60"],
            "codigo_61": ["CÓDIGO 61", "CODIGO 61", "COD61"],
            "documento": ["DOCUMENTO", "DOCUMENTOS"],
        }
        for key, aliases in _aliases.items():
            if not row_data.get(key):
                for alias in aliases:
                    col_idx = _upper.get(alias.upper())
                    if col_idx:
                        row_data[key] = ws.cell(row=row_idx, column=col_idx).value
                        break

        # Atributos técnicos: Atrib_N_Valor ou VALOR_N
        attrs = []
        for n in range(1, MAX_ATTRIBUTES + 1):
            col_idx = headers.get(f"Atrib_{n}_Valor") or _upper.get(f"VALOR_{n}")
            val = ws.cell(row=row_idx, column=col_idx).value if col_idx else None
            attrs.append(val)
        row_data["attributes"] = attrs

        # Ignorar linhas completamente vazias
        if row_data.get("sin") is None:
            continue

        items.append(row_data)

    log.info(f"Total de itens lidos: {len(items)}")
    return wb, items


def color_row(wb: openpyxl.Workbook, row: int, status: str) -> None:
    """Pinta a célula A da linha com a cor correspondente ao status."""
    ws = wb.active
    fill = {
        "ok": FILL_GREEN,
        "error": FILL_RED,
        "duplicate": FILL_ORANGE,
        "skipped": FILL_ORANGE,
        "needs_review": FILL_YELLOW,
    }.get(status, FILL_RED)
    ws.cell(row=row, column=1).fill = fill


def save_excel(wb: openpyxl.Workbook, path: Path | None = None) -> None:
    """Salva o workbook (com as cores atualizadas)."""
    path = path or EXCEL_PATH
    wb.save(path)
    log.debug(f"Excel salvo: {path}")


def enrich_missing_data(items: list[dict]) -> list[dict]:
    """Preenche campos vazios usando dados de itens vizinhos na planilha.

    A planilha tem sequência lógica — itens do mesmo equipamento vêm juntos.
    Quando um campo está vazio, busca no item anterior ou seguinte com mesmo grupo.
    """
    def _get_neighbor_value(items, idx, field, match_field):
        """Busca valor de 'field' no vizinho anterior ou seguinte que compartilhe 'match_field'."""
        current = items[idx]
        current_match = current.get(match_field) if match_field else None

        # Vizinho anterior
        if idx > 0:
            prev = items[idx - 1]
            if prev.get(field):
                if match_field is None or str(prev.get(match_field, "")) == str(current_match or ""):
                    return prev.get(field), str(prev.get("sin", "?"))
        # Vizinho seguinte
        if idx < len(items) - 1:
            nxt = items[idx + 1]
            if nxt.get(field):
                if match_field is None or str(nxt.get(match_field, "")) == str(current_match or ""):
                    return nxt.get(field), str(nxt.get("sin", "?"))
        return None, None

    enriched_count = 0

    for idx, item in enumerate(items):
        sin = str(item.get("sin", "?"))
        item.setdefault("_inferred", {})

        # Empresa
        if not item.get("empresa"):
            value, source_sin = None, None
            # Tentar vizinho com mesmo NCM
            if item.get("ncm"):
                value, source_sin = _get_neighbor_value(items, idx, "empresa", "ncm")
            # Tentar vizinho com mesmo UNSPSC
            if not value and item.get("unspsc"):
                value, source_sin = _get_neighbor_value(items, idx, "empresa", "unspsc")
            # Fallback: Part Number com prefixo ISK = BAKER HUGHES
            if not value and item.get("part_number") and str(item["part_number"]).upper().startswith("ISK"):
                value = "BAKER HUGHES"
                source_sin = "regra ISK"
            if value:
                item["empresa"] = value
                item["_inferred"]["empresa"] = f"{value} (de {source_sin})"
                log.warning(f"SIN {sin}: empresa vazia, inferido '{value}' do item vizinho (SIN {source_sin})")
                enriched_count += 1

        # NCM
        if not item.get("ncm"):
            value, source_sin = None, None
            if item.get("unspsc"):
                value, source_sin = _get_neighbor_value(items, idx, "ncm", "unspsc")
            if value:
                item["ncm"] = value
                item["_inferred"]["ncm"] = f"{value} (de SIN {source_sin})"
                log.warning(f"SIN {sin}: ncm vazio, inferido '{value}' do item vizinho (SIN {source_sin})")
                enriched_count += 1

        # UNSPSC
        if not item.get("unspsc"):
            value, source_sin = None, None
            if item.get("ncm"):
                value, source_sin = _get_neighbor_value(items, idx, "unspsc", "ncm")
            if value:
                item["unspsc"] = value
                item["_inferred"]["unspsc"] = f"{value} (de SIN {source_sin})"
                log.warning(f"SIN {sin}: unspsc vazio, inferido '{value}' do item vizinho (SIN {source_sin})")
                enriched_count += 1

        # codigo_60 — não pode inferir, apenas avisar
        if not item.get("codigo_60"):
            log.warning(f"SIN {sin}: codigo_60 vazio — step Relacionamento será pulado")

    if enriched_count:
        log.info(f"Enriquecimento: {enriched_count} campo(s) preenchido(s) via vizinhos")

    return items


def validate_documents(items: list[dict]) -> list[dict]:
    """Valida se os arquivos de documento existem no disco antes de processar.

    Retorna a lista de itens com campo '_missing_docs' preenchido.
    """
    for item in items:
        doc_str = item.get("documento")
        if not doc_str:
            item["_doc_files"] = []
            item["_missing_docs"] = []
            continue

        doc_names = [d.strip() for d in str(doc_str).split(";") if d.strip()]
        resolved = []
        missing = []
        for d in doc_names:
            full_path = DOCUMENTS_DIR / d
            if full_path.exists():
                resolved.append(str(full_path))
            else:
                # Tentar encontrar com extensão (ex: nome sem .pdf)
                matches = list(DOCUMENTS_DIR.glob(f"{d}.*"))
                if matches:
                    resolved.append(str(matches[0]))
                else:
                    missing.append(d)
        item["_doc_files"] = resolved
        item["_missing_docs"] = missing

        if missing:
            log.warning(
                f"SIN {item.get('sin')}: documentos não encontrados: {missing}"
            )

    return items
