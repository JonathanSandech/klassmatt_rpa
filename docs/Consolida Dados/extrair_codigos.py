"""
Extrai SIN, Código 61, Código 60, NCM e UNSPSC da planilha Acompanhamento.

Estratégia para NCM/UNSPSC:
1. Lookup direto: SINs já classificados na PARA ANALISE 2 (CONFIANÇA=DIRETO)
2. Match exato: equipamento idêntico ao da PA2 (CONFIANÇA=ALTA/MÉDIA)
3. Match por keywords: palavras-chave do equipamento (CONFIANÇA=KEYWORD)
4. Sem match: equipamento não encontrado (CONFIANÇA=BAIXA)
5. Validação: cruza contra listas de códigos válidos do Klassmatt
"""

import re
from pathlib import Path
from collections import Counter

import openpyxl

SCRIPT_DIR = Path(__file__).parent
ACOMPANHAMENTO = SCRIPT_DIR / "Acompanhamento_SIN-060813006.xlsx"
PARA_ANALISE_2 = SCRIPT_DIR / "PARA ANALISE 2.xlsx"
NCM_KLASSMATT = SCRIPT_DIR / "ncm_klassmatt.xlsx"
UNSPSC_KLASSMATT = SCRIPT_DIR / "unspsc_klassmatt.xlsx"
OUTPUT_FILE = SCRIPT_DIR / "output_codigos.xlsx"

# Acompanhamento: sheet "Lista", header row 4, data row 5+
ACOMP_SHEET = "Lista"
ACOMP_DATA_START = 5
ACOMP_COL_SIN = 1
ACOMP_COL_DESC = 4

# Padrões para extrair códigos 61/60 da descrição
CODE_PATTERNS = [
    re.compile(r"(61\d{10})-(60\d{10})"),
    re.compile(r"(\d{12})-(\d{12})"),
    re.compile(r"(61\d{10})-"),
]

# Padrão para encontrar onde começam os códigos na descrição
CODES_START = re.compile(r"\s+(?:61\d{10}|62\d{10}|\d{12})-")

# Padrão para remover Part Number + Empresa do final do nome do equipamento
PN_EMPRESA_TAIL = re.compile(
    r"\s+[A-Z]{1,3}[A-Z0-9]{5,}\s+"
    r"(?:BAKER HUGHES|BAKER HUG|BAKER H|BAKER|BAKE|BH|VELAN|DALIAN\s?\w*)\s*$"
)

# Palavras genéricas demais para keyword matching
STOP_WORDS = {
    "AND", "FOR", "THE", "WITH", "SET", "TYPE", "SIZE", "KIT", "COMPLETE",
    "STANDARD", "ASSY", "ASSEMBLY", "MODEL", "SPEC", "SPECIFICATION",
    "PER", "REPAIR", "SPARE", "PART", "ITEM", "UNIT", "PIECE",
}


def extract_codes(desc: str) -> tuple[str, str]:
    """Retorna (codigo_61, codigo_60) extraídos da descrição."""
    if not desc:
        return "", ""
    for pattern in CODE_PATTERNS:
        m = pattern.search(desc)
        if m:
            cod_61 = m.group(1)
            cod_60 = m.group(2) if m.lastindex >= 2 else ""
            return cod_61, cod_60
    return "", ""


def extract_equipment(desc: str) -> str:
    """Extrai nome do equipamento da SIN_Descrição do Acompanhamento."""
    if not desc:
        return ""
    m = CODES_START.search(desc)
    if m:
        equip = desc[: m.start()].strip()
    else:
        equip = desc[:50].strip()
    equip = PN_EMPRESA_TAIL.sub("", equip).strip()
    return equip.upper()


def load_para_analise_2() -> dict:
    """Carrega PARA ANALISE 2 -> {SIN: {ncm, unspsc, equip}}."""
    print(f"  Carregando {PARA_ANALISE_2.name}...")
    wb = openpyxl.load_workbook(PARA_ANALISE_2, data_only=True)
    ws = wb["Sheet1"]
    data = {}
    for r in range(2, ws.max_row + 1):
        sin = ws.cell(r, 1).value
        equip = ws.cell(r, 2).value
        ncm = ws.cell(r, 7).value
        unspsc = ws.cell(r, 8).value
        if sin and ncm:
            data[sin] = {
                "ncm": str(ncm).strip(),
                "unspsc": str(int(unspsc)) if unspsc else "",
                "equip": str(equip).strip().upper() if equip else "",
            }
    print(f"    {len(data)} SINs com NCM/UNSPSC")
    return data


def load_valid_ncms() -> set[str]:
    """Carrega NCMs válidos do Klassmatt."""
    print(f"  Carregando {NCM_KLASSMATT.name}...")
    wb = openpyxl.load_workbook(NCM_KLASSMATT, data_only=True)
    ws = wb["Lista"]
    ncms = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v:
            ncms.add(str(v).strip())
    print(f"    {len(ncms)} NCMs validos")
    return ncms


def load_valid_unspscs() -> set[str]:
    """Carrega UNSPSCs válidos do Klassmatt (Commodity + Class codes)."""
    print(f"  Carregando {UNSPSC_KLASSMATT.name}...")
    wb = openpyxl.load_workbook(UNSPSC_KLASSMATT, data_only=True)
    ws = wb["UNSPSC"]
    codes = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v:
            codes.add(str(int(v)))
        cl = ws.cell(r, 3).value
        if cl:
            codes.add(str(int(cl)))
    print(f"    {len(codes)} UNSPSCs validos (commodity + class)")
    return codes


def build_inference_table(pa2_data: dict) -> dict:
    """Constrói tabela: equipamento -> {ncm_counter, unspsc_counter}."""
    table = {}
    for info in pa2_data.values():
        equip = info["equip"]
        if not equip:
            continue
        if equip not in table:
            table[equip] = {"ncm": Counter(), "unspsc": Counter()}
        table[equip]["ncm"][info["ncm"]] += 1
        if info["unspsc"]:
            table[equip]["unspsc"][info["unspsc"]] += 1
    return table


def build_keyword_index(inference_table: dict) -> dict:
    """Constrói índice: keyword -> {ncm_counter, unspsc_counter}.

    Agrega NCM/UNSPSC de todos equipamentos que contêm a keyword.
    """
    index = {}
    for equip, data in inference_table.items():
        words = set()
        for w in equip.split():
            if len(w) > 2 and w not in STOP_WORDS and not re.match(r"^\d", w):
                words.add(w)
        top_ncm = data["ncm"].most_common(1)[0][0]
        top_unspsc = data["unspsc"].most_common(1)[0][0] if data["unspsc"] else ""
        weight = sum(data["ncm"].values())
        for word in words:
            if word not in index:
                index[word] = {"ncm": Counter(), "unspsc": Counter()}
            index[word]["ncm"][top_ncm] += weight
            if top_unspsc:
                index[word]["unspsc"][top_unspsc] += weight
    return index


def _best_from_counter(counter: Counter) -> tuple[str, float]:
    """Retorna (valor_mais_frequente, percentual)."""
    if not counter:
        return "", 0.0
    top, top_count = counter.most_common(1)[0]
    total = sum(counter.values())
    return top, top_count / total


def infer_exact(
    equip: str, table: dict, valid_ncms: set, valid_unspscs: set
) -> tuple[str, str, str]:
    """Infere NCM/UNSPSC por match exato de equipamento."""
    if equip not in table:
        return "", "", ""

    entry = table[equip]
    ncm_top, ncm_pct = _best_from_counter(entry["ncm"])
    unspsc_top, _ = _best_from_counter(entry["unspsc"])

    if len(entry["ncm"]) == 1:
        confianca = "ALTA"
    elif ncm_pct > 0.7:
        confianca = "MEDIA"
    else:
        confianca = "BAIXA"

    ncm = ncm_top if ncm_top in valid_ncms else ""
    unspsc = unspsc_top if unspsc_top in valid_unspscs else ""
    return ncm, unspsc, confianca


def infer_keyword(
    equip: str, kw_index: dict, valid_ncms: set, valid_unspscs: set
) -> tuple[str, str, str]:
    """Infere NCM/UNSPSC por keywords do equipamento."""
    words = [
        w for w in equip.split()
        if len(w) > 2 and w not in STOP_WORDS and not re.match(r"^\d", w)
    ]
    if not words:
        return "", "", ""

    # Encontrar keywords que existem no índice
    matched = [w for w in words if w in kw_index]
    if not matched:
        return "", "", ""

    # Agregar votos de todas as keywords encontradas
    ncm_votes = Counter()
    unspsc_votes = Counter()
    for kw in matched:
        for ncm, count in kw_index[kw]["ncm"].items():
            ncm_votes[ncm] += count
        for unspsc, count in kw_index[kw]["unspsc"].items():
            unspsc_votes[unspsc] += count

    ncm_top, _ = _best_from_counter(ncm_votes)
    unspsc_top, _ = _best_from_counter(unspsc_votes)

    ncm = ncm_top if ncm_top in valid_ncms else ""
    unspsc = unspsc_top if unspsc_top in valid_unspscs else ""

    if ncm:
        return ncm, unspsc, "KEYWORD"
    return "", "", ""


def main():
    print("=" * 60)
    print("Extrair Codigos + NCM/UNSPSC")
    print("=" * 60)

    # Passo 1: Carregar referências
    print("\n[1/5] Carregando dados de referencia...")
    pa2_data = load_para_analise_2()
    valid_ncms = load_valid_ncms()
    valid_unspscs = load_valid_unspscs()

    # Passo 2: Construir tabelas de inferência
    print("\n[2/5] Construindo tabelas de inferencia...")
    inference_table = build_inference_table(pa2_data)
    kw_index = build_keyword_index(inference_table)
    n_1to1 = sum(1 for e in inference_table.values() if len(e["ncm"]) == 1)
    print(f"    {len(inference_table)} equipamentos unicos (match exato)")
    print(f"    {n_1to1} com mapeamento 1:1 NCM")
    print(f"    {len(kw_index)} keywords unicas (match fuzzy)")

    # Passo 3: Processar Acompanhamento
    print(f"\n[3/5] Processando {ACOMPANHAMENTO.name}...")
    wb = openpyxl.load_workbook(ACOMPANHAMENTO, data_only=True)
    ws = wb[ACOMP_SHEET]

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Codigos"
    ws_out.append(
        ["SIN", "CODIGO 61", "CODIGO 60", "NCM", "UNSPSC", "CONFIANCA", "EQUIPAMENTO"]
    )

    stats = Counter()

    for row_idx in range(ACOMP_DATA_START, ws.max_row + 1):
        id_sin = ws.cell(row_idx, ACOMP_COL_SIN).value
        if not id_sin:
            continue

        desc = str(ws.cell(row_idx, ACOMP_COL_DESC).value or "")
        cod_61, cod_60 = extract_codes(desc)
        equip = extract_equipment(desc)

        # Estratégia em cascata
        if id_sin in pa2_data:
            # 1. Lookup direto
            info = pa2_data[id_sin]
            ncm = info["ncm"] if info["ncm"] in valid_ncms else ""
            unspsc = info["unspsc"] if info["unspsc"] in valid_unspscs else ""
            confianca = "DIRETO"
            equip_out = info["equip"] or equip
        else:
            equip_out = equip
            # 2. Match exato por equipamento
            ncm, unspsc, confianca = infer_exact(
                equip, inference_table, valid_ncms, valid_unspscs
            )
            if not ncm:
                # 3. Match por keywords
                ncm, unspsc, confianca = infer_keyword(
                    equip, kw_index, valid_ncms, valid_unspscs
                )
            if not ncm:
                # 4. Sem match
                ncm, unspsc, confianca = "", "", "BAIXA"

        stats[confianca] += 1
        ws_out.append([id_sin, cod_61, cod_60, ncm, unspsc, confianca, equip_out])

    # Passo 4: Salvar
    print(f"\n[4/5] Salvando {OUTPUT_FILE.name}...")
    wb_out.save(OUTPUT_FILE)

    # Passo 5: Relatório
    total = sum(stats.values())
    print(f"\n{'=' * 60}")
    print(f"RESULTADO: {total} SINs processados")
    print(f"{'=' * 60}")
    com_ncm = 0
    for nivel in ["DIRETO", "ALTA", "MEDIA", "KEYWORD", "BAIXA"]:
        count = stats.get(nivel, 0)
        pct = count / total * 100 if total else 0
        bar = "#" * int(pct / 2)
        print(f"  {nivel:10s}: {count:5d} ({pct:5.1f}%) {bar}")
        if nivel != "BAIXA":
            com_ncm += count
    print(f"  {'TOTAL':10s}: {total:5d}")
    print(f"\n  Com NCM preenchido: {com_ncm} ({com_ncm/total*100:.1f}%)")
    print(f"  Sem NCM (revisao manual): {total - com_ncm} ({(total-com_ncm)/total*100:.1f}%)")


if __name__ == "__main__":
    main()
